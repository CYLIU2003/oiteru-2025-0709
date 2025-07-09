import mysql.connector  # MySQLデータベースと接続するためのライブラリをインポート
from mysql.connector import Error  # エラー処理のためにErrorクラスをインポート
import pandas as pd  # データを扱うためのpandasライブラリをインポート
import os  # オペレーティングシステムの機能を使用するためのosモジュールをインポート
import sys  # Pythonのシステム関連の機能を使用するためのsysモジュールをインポート
import binascii  # バイナリデータとASCII文字列を相互に変換するためのbinasciiモジュールをインポート
import nfc  # NFC通信に関連するライブラリをインポート
import ndef
import datetime  # 日時を扱うためのdatetimeモジュールをインポート
import time  # 時間を扱うためのtimeモジュールをインポート
from functools import partial  # 部分適用を行うためのpartial関数をインポート
import openpyxl  # Excelファイルを扱うためのopenpyxlライブラリをインポート
from nfc.clf import rcs380  # NFC通信のための特定のクラスrcs380をインポート
import shutil  # ★ 旧版 dayupdate / make_backup 用

# パスを設定（親ディレクトリをsys.pathに追加）
sys.path.insert(1, os.path.split(sys.path[0])[0])

# サービスコードと現在の年を取得
service_code = 0x090f  # サービスコードを定義
dt = datetime.datetime.now()  # 現在の日付と時刻を取得
freshmen = dt.year - 2000  # 現在の年から2000を引いて新入生の年を計算
allow_list = []  # 許可リストを初期化

# 許可リストを生成
for i in range(10):  # 0から9までのループで
    allow_list.append(f'g{str(freshmen - i)}')  # 新入生の年からiを引いた文字列を追加

# サーバー接続を確立する関数
def create_server_connection(host_name, user_name, user_password, dbname):
    connection = None  # 接続を初期化
    try:
        connection = mysql.connector.connect(  # MySQLデータベースへの接続を試みる
            host=host_name,
            user=user_name,
            passwd=user_password,
            database=dbname,
            auth_plugin="mysql_native_password"  # 認証プラグインを指定
        )
    except Error as err:  # エラーが発生した場合
        print(f"Error: '{err}'")  # エラーメッセージを表示
    return connection  # 接続オブジェクトを返す

# サーバー接続を確立する関数（データベースなし）
def create_server_connection2(host_name, user_name, user_password):
    connection = None  # 接続を初期化
    try:
        connection = mysql.connector.connect(  # MySQLデータベースへの接続を試みる
            host=host_name,
            user=user_name,
            passwd=user_password,
            auth_plugin="mysql_native_password"  # 認証プラグインを指定
        )
    except Error as err:  # エラーが発生した場合
        print(f"Error: '{err}'")  # エラーメッセージを表示
    return connection  # 接続オブジェクトを返す

# データベースを作成する関数
def create_database_if_not_exists():
    conn = create_server_connection2("localhost", "root", "Hiramekigo@1")  # データベースなしで接続
    if conn is not None:
        cur = conn.cursor()  # カーソルを作成
        cur.execute("CREATE DATABASE IF NOT EXISTS userdb")  # データベースを作成
        conn.commit()  # 変更をコミット
        cur.close()  # カーソルを閉じる
        conn.close()  # 接続を閉じる
        print("データベース 'userdb' が作成されました。")

# 初期情報を作成する関数
def make_info():
    conn = create_server_connection("localhost", "root", "Hiramekigo@1", "userdb")  # データベースに接続
    cur = conn.cursor(buffered=True)  # カーソルを作成
    cur.execute('DROP TABLE IF EXISTS info') # 既存のテーブルを削除
    cur.execute('CREATE TABLE IF NOT EXISTS info(pass VARCHAR(50) PRIMARY KEY, daycount INT, updatecount INT, freq INT, maximum INT, done INT, list VARCHAR(50))') # 新しいテーブルを作成
    cur.execute('INSERT INTO info (pass, daycount, updatecount, freq, maximum, done, list) VALUES("test1", 3, 0, 4, 2, 0, "hello")') # 初期データを挿入
    conn.commit() # 変更をコミット
    cur.close() # カーソルを閉じる
    conn.close() # 接続を閉じる

# ユーザー情報を作成する関数
def make_user():
    conn = create_server_connection("localhost", "root", "Hiramekigo@1", "userdb")  # データベースに接続
    cur = conn.cursor(buffered=True)  # カーソルを作成
    cur.execute('CREATE TABLE IF NOT EXISTS user(id INT PRIMARY KEY, cardid VARCHAR(50), allow INT, entry DATETIME, stock INT, today INT, total INT, last1 DATETIME, last2 DATETIME, last3 DATETIME, last4 DATETIME, last5 DATETIME, last6 DATETIME, last7 DATETIME, last8 DATETIME, last9 DATETIME, last10 DATETIME)') # 新しいユーザーテーブルを作成

    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # 現在の日付と時刻を取得（文字列形式）
    try:
        cur.execute(f'INSERT INTO user (id, cardid, allow, entry, stock, today, total, last1, last2, last3, last4, last5, last6, last7, last8, last9, last10) VALUES(0, "card121", 1, "{now}", 1, 1, 1, "{now}", "{now}", "{now}", "{now}", "{now}", "{now}", "{now}", "{now}", "{now}", "{now}")') # 初期ユーザーデータを挿入
    except:
        pass
    conn.commit() # 変更をコミット
    cur.close() # カーソルを閉じる
    conn.close() # 接続を閉じる

# ユニット情報を作成する関数
def make_unit(unit_name, unit_pass, unit_stock, unit_available):
    cu = call_units() # 既存のユニット情報を取得
    conn = create_server_connection("localhost", "root", "Hiramekigo@1", "userdb") # データベースに接続
    cur = conn.cursor(buffered=True) # カーソルを作成
    cur.execute('CREATE TABLE IF NOT EXISTS units(id INT PRIMARY KEY AUTO_INCREMENT, name VARCHAR(50), pass VARCHAR(50), stock INT, connect INT, available INT)') # 新しいユニットテーブルを作成

    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S") # 現在の日付と時刻を取得（文字列形式）
    cur.execute(f'CREATE USER "{unit_name}"@"%" IDENTIFIED WITH mysql_native_password BY "{unit_pass}"') # 新しいユーザーを作成
    cur.execute(f'GRANT ALL PRIVILEGES ON userdb.* TO "{unit_name}"@"%"') # ユーザーにすべての権限を付与
    cur.execute(f'INSERT INTO units (id, name, pass, stock, connect, available) VALUES({len(cu)+1}, "{unit_name}", "{unit_pass}", {unit_stock}, 1, {unit_available})')    # 新しいユニットをテーブルに挿入
    cur.execute(f'UPDATE info SET list = "{len(cu)+1}:{unit_name}"') # infoテーブルのリストを更新

    conn.commit() # 変更をコミット
    cur.close() # カーソルを閉じる
    conn.close() # 接続を閉じる

# ヒストリー情報を作成する関数
def make_his():
    conn = create_server_connection("localhost", "root", "Hiramekigo@1", "userdb") # データベースに接続
    cur = conn.cursor(buffered=True) # カーソルを作成
    try:
        call_his() # ヒストリー情報を取得
    except: # 取得に失敗した場合
        cur.execute('CREATE TABLE IF NOT EXISTS his(id INT PRIMARY KEY AUTO_INCREMENT, txt VARCHAR(50))') # ヒストリー用テーブルを作成

        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S") # 現在の日付と時刻を取得
        cur.execute(f'INSERT INTO his (txt) VALUES("{now}:sample")')    # 初期ヒストリーデータを挿入
    conn.commit() # 変更をコミット
    cur.close() # カーソルを閉じる
    conn.close() # 接続を閉じる

# ユニット情報を取得する関数
def call_units():
    conn = create_server_connection("localhost", "root", "Hiramekigo@1", "userdb") # データベースに接続
    cur = conn.cursor(buffered=True) # カーソルを作成
    cur.execute('SELECT * FROM units ORDER BY id ASC') # ユニットテーブルの全データを取得
    get = cur.fetchall() # すべての行を取得
    cur.close() # カーソルを閉じる
    conn.close() # 接続を閉じる
    return get # 取得したデータを返す

# ヒストリー情報を取得する関数
def call_his():
    conn = create_server_connection("localhost", "root", "Hiramekigo@1", "userdb") # データベースに接続
    cur = conn.cursor(buffered=True) # カーソルを作成
    cur.execute('SELECT * FROM his ORDER BY id ASC') # ヒストリーテーブルの全データを取得
    get = cur.fetchall() # すべての行を取得
    cur.close() # カーソルを閉じる
    conn.close() # 接続を閉じる
    return get # 取得したデータを返す

# 情報を取得する関数
def call_info():
    conn = create_server_connection("localhost", "root", "Hiramekigo@1", "userdb") # データベースに接続
    cur = conn.cursor(buffered=True) # カーソルを作成
    cur.execute('SELECT * FROM info ORDER BY pass ASC') # infoテーブルの全データをパス順に取得
    get = cur.fetchone() # 一行だけ取得
    cur.close() # カーソルを閉じる
    conn.close() # 接続を閉じる
    return get # 取得したデータを返す

# 情報を取得する関数（別のサーバー用）
def call_info2():
    conn = create_server_connection("192.168.11.2", "hirameki2", "tcu@2122","userdb") # 別サーバーに接続
    cur = conn.cursor(buffered=True) # カーソルを作成
    cur.execute('SELECT * FROM info ORDER BY pass ASC') # infoテーブルの全データをパス順に取得
    get = cur.fetchone() # 一行だけ取得
    cur.close() # カーソルを閉じる
    conn.close() # 接続を閉じる
    return get # 取得したデータを返す

# 情報を更新する関数
def update_info(num, data):
    data_list = ['pass', 'daycount', 'updatecount', 'freq', 'maximum', 'done', 'list'] # 更新するカラム名のリスト
    call_data = data_list[num] # 更新するカラム名を特定
    conn = create_server_connection("localhost", "root", "Hiramekigo@1", "userdb") # データベースに接続
    cur = conn.cursor(buffered=True) # カーソルを作成
    cur.execute(f'UPDATE info SET {call_data} = "{data}"') # 指定したカラムを更新
    conn.commit() # 変更をコミット
    cur.execute('SELECT * FROM info ORDER BY pass ASC') # 更新後のデータをパス順に取得
    get = cur.fetchone() # 一行だけ取得
    cur.close() # カーソルを閉じる
    conn.close() # 接続を閉じる

# ユーザー情報を更新する関数
def update_user(idnum, num, data):
    data_list = ['id', 'cardid', 'allow', 'entry', 'stock', 'today', 'total', 'last1', 'last2', 'last3', 'last4', 'last5', 'last6', 'last7', 'last8', 'last9', 'last10'] # 更新するカラム名のリスト
    call_data = data_list[num] # 更新するカラム名を特定
    conn = create_server_connection("localhost", "root", "Hiramekigo@1", "userdb") # データベースに接続
    cur = conn.cursor(buffered=True) # カーソルを作成
    cur.execute(f'UPDATE user SET {call_data} = "{data}" WHERE id = {idnum}') # 指定したユーザーのカラムを更新
    conn.commit() # 変更をコミット
    cur.execute('SELECT * FROM user ORDER BY id ASC') # 更新後のデータをID順に取得
    get = cur.fetchall() # 全ての行を取得
    cur.close() # カーソルを閉じる
    conn.close() # 接続を閉じる

# ユニット情報を更新する関数
def update_unit(idnum, num, data):
    data_list = ['id', 'name', 'pass', 'stock', 'connect', 'available'] # 更新するカラム名のリスト
    call_data = data_list[num] # 更新するカラム名を特定
    conn = create_server_connection("localhost", "root", "Hiramekigo@1", "userdb") # データベースに接続
    cur = conn.cursor(buffered=True) # カーソルを作成
    cur.execute(f'UPDATE units SET {call_data} = "{data}" WHERE id = {idnum}') # 指定したユニットのカラムを更新
    conn.commit() # 変更をコミット
    cur.execute('SELECT * FROM units ORDER BY id ASC') # 更新後のデータをID順に取得
    get = cur.fetchall() # 全ての行を取得
    cur.close() # カーソルを閉じる
    conn.close() # 接続を閉じる

# ユーザー情報を更新する関数（別のサーバー用）
def update_user2(idnum, num, data):
    data_list = ['id', 'cardid', 'allow', 'entry', 'stock', 'today', 'total', 'last1', 'last2', 'last3', 'last4', 'last5', 'last6', 'last7', 'last8', 'last9', 'last10'] # 更新するカラム名のリスト
    call_data = data_list[num] # 更新するカラム名を特定
    conn = create_server_connection("192.168.11.2", "hirameki2", "tcu@2122","userdb") # 別サーバーに接続
    cur = conn.cursor(buffered=True) # カーソルを作成
    cur.execute(f'UPDATE user SET {call_data} = "{data}" WHERE id = {idnum}') # 指定したユーザーのカラムを更新
    conn.commit() # 変更をコミット
    cur.execute('SELECT * FROM user ORDER BY id ASC') # 更新後のデータをID順に取得
    get = cur.fetchall() # 全ての行を取得
    cur.close() # カーソルを閉じる
    conn.close() # 接続を閉じる

# ユーザーを削除する関数
def delete_user(id):
    conn = create_server_connection("localhost", "root", "Hiramekigo@1", "userdb") # データベースに接続
    cur = conn.cursor(buffered=True) # カーソルを作成
    cur.execute(f'DELETE FROM user WHERE id={id}') # 指定したIDのユーザーを削除
    print("消したよ") # 削除完了のメッセージを表示
    conn.commit() # 変更をコミット
    cur.close() # カーソルを閉じる
    conn.close() # 接続を閉じる

# ユニットを削除する関数
def delete_unit(id):
    conn = create_server_connection("localhost", "root", "Hiramekigo@1", "userdb") # データベースに接続
    cur = conn.cursor(buffered=True) # カーソルを作成
    cur.execute(f'DELETE FROM units WHERE id={id}') # 指定したIDのユニットを削除
    print("消したよ") # 削除完了のメッセージを表示
    conn.commit() # 変更をコミット
    cur.close() # カーソルを閉じる
    conn.close() # 接続を閉じる

# ユーザー情報を取得する関数
def call_user():
    conn = create_server_connection("localhost", "root", "Hiramekigo@1", "userdb") # データベースに接続
    cur = conn.cursor(buffered=True) # カーソルを作成
    cur.execute('SELECT * FROM user ORDER BY id ASC') # ユーザーテーブルの全データを取得
    get = cur.fetchall() # すべての行を取得
    cur.close() # カーソルを閉じる
    conn.close() # 接続を閉じる
    return get # 取得したデータを返す

# 別のサーバーからユーザー情報を取得する関数
def call_user2():
    conn = create_server_connection("192.168.11.2", "hirameki2", "tcu@2122","userdb") # 別サーバーに接続
    cur = conn.cursor(buffered=True) # カーソルを作成
    cur.execute('SELECT * FROM user ORDER BY id ASC') # ユーザーテーブルの全データを取得
    get = cur.fetchall() # すべての行を取得
    cur.close() # カーソルを閉じる
    conn.close() # 接続を閉じる
    return get # 取得したデータを返す

# データベースの初期設定を行う関数
def set_up():
    """初回起動時に DB / テーブル / 初期レコードを生成する統合関数。"""
    # 1) データベースが無ければ作成
    create_database_if_not_exists()

    # 2) 基本テーブル
    make_info()
    make_user()
    make_his()

    # 3) バックアップディレクトリ確保
    os.makedirs("backup", exist_ok=True)

    # 4) 毎日 1 回だけ info.updatecount を進め、週次でバックアップ   旧 dayupdate()
    try:
        ci = call_info()
        today = datetime.date.today()
        last_day = datetime.date.fromtimestamp(
            os.path.getmtime("backup/info_last.bak")
        ) if os.path.exists("backup/info_last.bak") else None

        # 日付が変わっていればカウントを進める
        if last_day is None or today != last_day:
            update_info(2, ci[2] + 1)          # updatecount++
            shutil.copyfile("backup/info_last.bak",
                            f"backup/info_{today}.bak") if last_day else None
            with open("backup/info_last.bak", "wb") as f:
                f.write(b"")                   # タイムスタンプ更新のみ
    except Exception as e:
        print(f"[WARN] 日次更新 / バックアップに失敗: {e}")


# ユーザー情報を整合させる関数
def alignment_user():
    cu = list(call_user()) # 現在のユーザー情報を取得
    conn = create_server_connection("localhost", "root", "Hiramekigo@1", "userdb") # データベースに接続
    cur = conn.cursor(buffered=True) # カーソルを作成
    cur.execute('TRUNCATE user') # ユーザーテーブルを空にする
    conn.commit()

    for i in range(len(cu)): # ユーザーのリストをループ
        cur.execute(f'INSERT INTO user (id, cardid, allow, entry, stock, today, total, last1, last2, last3, last4, last5, last6, last7, last8, last9, last10) VALUES({i}, "{cu[i][1]}", {cu[i][2]}, "{cu[i][3]}", {cu[i][4]}, {cu[i][5]}, {cu[i][6]}, "{cu[i][7]}", "{cu[i][8]}", "{cu[i][9]}", "{cu[i][10]}", "{cu[i][11]}", "{cu[i][12]}", "{cu[i][13]}", "{cu[i][14]}", "{cu[i][15]}", "{cu[i][16]}")') # ユーザーデータを再挿入
    conn.commit() # 変更をコミット
    cur.close() # カーソルを閉じる
    conn.close() # 接続を閉じる

# ユニット情報を整合させる関数
def alignment_units():
    cu = list(call_units()) # 現在のユニット情報を取得
    conn = create_server_connection("localhost", "root", "Hiramekigo@1", "userdb") # データベースに接続
    cur = conn.cursor(buffered=True) # カーソルを作成
    cur.execute('TRUNCATE units') # ユニットテーブルを空にする
    conn.commit()

    for i in range(len(cu)): # ユニットのリストをループ
        cur.execute(f'INSERT INTO units (id, name, pass, stock, connect, available) VALUES({i+1}, "{cu[i][1]}", "{cu[i][2]}", {cu[i][3]}, {cu[i][4]}, {cu[i][5]})') # ユニットデータを再挿入
    conn.commit() # 変更をコミット
    cur.close() # カーソルを閉じる
    conn.close() # 接続を閉じる

# 日付を更新する関数
def dayupdate():
    dt = datetime.datetime.now()  # 現在の日付と時刻を取得
    freshmen = dt.year - 2000  # 現在の年から2000を引いて新入生の年を計算
    allow_list = []  # 許可リストを初期化
    
    for i in range(10):  # 0から9までのループで
        allow_list.append(f'g{str(freshmen - i)}')  # 新入生の年からiを引いた文字列を追加
    
    conn = create_server_connection("localhost", "root", "Hiramekigo@1", "userdb")  # データベースに接続
    cur = conn.cursor(buffered=True)  # カーソルを作成
    today = dt.day  # 今日の日にちを取得
    new_flag = False  # 新しい日付フラグを初期化
    plus_flag = False  # 利用可能回数更新フラグを初期化
    ci = call_info()  # infoテーブルからデータを取得
    
    # ciが不正な場合は処理を終了
    if ci is None or len(ci) <= 4:
        cur.close()
        conn.close()
        return
    
    cu = call_user()  # userテーブルからデータを取得
    date = ci[1]  # infoテーブルの取得データから日付を取得
    newcount = ci[2]  # infoテーブルの取得データから更新カウントを取得
    
    print('new day?')  # 新しい日付かどうか確認
    if (today > date) or (date - today) > 10:  # 今日の日付が取得日付より新しいか、日付の差が10日以上
        print('日付が更新されました')  # 日付が更新されたことを表示
        update_info(1, today)  # infoテーブルの日付を更新
        update_info(2, newcount + 1)  # 更新カウントを1増やして更新
        ci = call_info()  # 更新後のinfoテーブルからデータを再取得
        new_flag = True  # 新しい日付フラグをTrueに設定
        
    if ci and len(ci) > 3 and ci[2] >= ci[3]:  # 更新カウントが最大値以上の場合
        print('利用可能回数を更新しました')  # 利用可能回数を更新することを表示
        plus_flag = True  # 利用可能フラグをTrueに設定
        update_info(2, 0)  # 利用可能カウントを0にリセット
        
    if new_flag:  # 新しい日付である場合
        print("実行")  # 実行メッセージを表示
        for tnum in range(len(cu)):  # すべてのユーザーについてループ
            getid = cu[tnum][1]  # ユーザーのIDを取得
            gettoday = cu[tnum][5]  # ユーザーの今日の利用数を取得
            getstock = cu[tnum][4]  # ユーザーの在庫を取得
            update_user(tnum, 5, 0)  # 今日の利用数を0にリセット
            if plus_flag and getstock < ci[4]:  # 在庫が最大を下回る場合
                update_user(tnum, 4, getstock + 1)  # 在庫を1増やす

# 完了フラグをゼロに戻す関数
def done_zero():
    conn = create_server_connection("localhost", "root", "Hiramekigo@1", "userdb")  # データベースに接続
    cur = conn.cursor(buffered=True)  # カーソルを作成
    ci = call_info()  # infoテーブルからデータを取得
    if ci is None or len(ci) <= 5:
        cur.close()
        conn.close()
        return False
    
    ci_flag = ci[5]  # infoテーブルから完了フラグを取得
    
    if ci_flag == 1:  # 完了フラグが1の場合
        cur.execute('UPDATE info SET done = 0')  # 完了フラグを0に設定
        conn.commit()  # 変更をコミット
        cur.close()  # カーソルを閉じる
        conn.close()  # 接続を閉じる
        return True  # 成功メッセージを返す
    else:
        cur.close()  # カーソルを閉じる
        conn.close()  # 接続を閉じる
        return False  # 成功しなかった場合

# 時間経過を確認する関数
def afrer(n, started):
    return time.time() - started > n  # 指定時間経過判定

# NFCタグが接続されたときの処理
def connected(tag):
    done_zero()  # 完了フラグをリセット
    allow = None  # 許可フラグを初期化
    ci = call_info()  # infoテーブルからデータを取得
    cu = call_user()  # userテーブルからデータを取得
    num = len(cu)  # ユーザー数を取得
    
    if num == 0:  # ユーザーが存在しない場合
        num = 1  # ユーザー数を1に設定
    
    dt = datetime.datetime.now()  # 現在の日付と時刻を取得
    ic_flag = False  # NFC許可フラグを初期化
    
    try:
        if isinstance(tag, nfc.tag.tt3.Type3Tag):  # タグがType3Tagのインスタンスの場合
            try:
                newdata = tag.idm.hex()  # NFCタグのIDを16進数形式で取得
                data = tag.dump()  # タグの情報をダンプ
                info = data[len(data) - 1]  # ダンプデータの最後の要素を取得
                
                for j in allow_list:  # 許可リストをループ
                    allow = j in info  # 許可リストに情報が含まれているか確認
                    if allow:  # 許可されている場合
                        ic_flag = True  # 許可フラグをTrueに設定
                        break  # ループを終了
                
                if ic_flag:  # 許可フラグがTrueの場合
                    print("allowed")
                    for t in range(num):  # ユーザー数分ループ
                        try:
                            getid = cu[t][1]  # ユーザーIDを取得
                        except:  # エラーが発生した場合
                            getid = None  # IDをNoneに設定
                            
                        if getid == newdata:  # 新しいデータがユーザーIDに一致する場合
                            conn = create_server_connection("localhost", "root", "Hiramekigo@1", "userdb")  # データベースに接続
                            cur = conn.cursor(buffered=True)  # カーソルを作成
                            cur.execute('UPDATE info SET done = 1')  # 完了フラグを1に設定
                            break  # ループを終了
                            
                        if t == len(cu) - 1 or getid is None:  # 最後のユーザーに達したか、IDがNoneの場合
                            print(cu)
                            conn = create_server_connection("localhost", "root", "Hiramekigo@1", "userdb")  # データベースに接続
                            cur = conn.cursor(buffered=True)  # カーソルを作成
                            now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # 現在の日付と時刻を取得（文字列形式）
                            cur.execute(f'INSERT INTO user (id, cardid, allow, entry, stock, today, total, last1, last2, last3, last4, last5, last6, last7, last8, last9, last10) VALUES({len(cu)}, "{newdata}", 1, "{now}", 2, 0, 0, "{now}", "{now}", "{now}", "{now}", "{now}", "{now}", "{now}", "{now}", "{now}", "{now}")')  # 新ユーザーを挿入
                            print('register_user')
                            cur.execute(f'UPDATE info SET list = "{len(cu)}:{newdata}"')  # infoテーブルのリストを更新
                    
                    conn.commit()  # 変更をコミット
                    cur.close()  # カーソルを閉じる
                    conn.close()  # 接続を閉じる
                
                return allow  # 許可状況を返す
            except Exception as e:  # エラーが発生した場合
                print("error: %s" % e)  # エラーメッセージを表示
                print("例外発生。係員にご問い合わせください")  # 注意メッセージを表示
                return False  # Falseを返す
        else:
            print("error: tag isn't Type3Tag")  # タグがType3Tagでない場合のエラーメッセージを表示
            return False  # Falseを返す
    except AttributeError:  # 属性エラーが発生した場合
        print('不正をするなバカもの')  # 警告メッセージを表示
        return False  # Falseを返す

# NFC接続を開始する関数
def do():
    started = time.time()  # 開始時間を記録
    clf = nfc.ContactlessFrontend('usb')  # NFCリーダーをUSBデバイスとして開く
    return clf.connect(rdwr={'on-connect': connected}, terminate=partial(afrer, 5, started))  # 接続処理を開始し、接続時の処理を指定

# 子機用の接続処理
def unit(tag):
    allow = None  # 許可フラグを初期化
    allow2 = False  # 子機許可フラグを初期化
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # 現在の日付と時刻を取得（文字列形式）
    ic_flag = False  # NFC許可フラグを初期化
    ci = call_info2()  # 別サーバーのinfoテーブルからデータを取得
    cu = call_user2()  # 別サーバーのuserテーブルからデータを取得
    
    try:
        if isinstance(tag, nfc.tag.tt3.Type3Tag):  # タグがType3Tagのインスタンスの場合
            try:
                data = tag.idm.hex()  # NFCタグのIDを16進数形式で取得
                for t in range(len(cu)):  # ユーザー数分ループ
                    getid = cu[t][1]  # ユーザーIDを取得
                    gettoday = cu[t][5]  # 今日の利用数を取得
                    getstock = cu[t][4]  # 残数を取得
                    gettotal = cu[t][6]  # 累計利用数を取得
                    allow = None  # 許可フラグを初期化
                    
                    if getid == data:  # タグIDがユーザーIDと一致する場合
                        allow2 = True  # 子機利用許可フラグをTrueに設定
                        if getstock > 0:  # 残数が0より大きい場合
                            allow = True  # 許可フラグをTrueに設定
                            update_user2(t, 5, gettoday + 1)  # 今日の利用数を1増やす
                            update_user2(t, 4, getstock - 1)  # 残数を1減らす
                            update_user2(t, 6, gettotal + 1)  # 累計利用数を1増やす
                            
                            # 直近の利用日時を更新
                            update_user2(t, 16, cu[t][15])  # last10に値を移動
                            update_user2(t, 15, cu[t][14])  # last9に値を移動
                            update_user2(t, 14, cu[t][13])  # last8に値を移動
                            update_user2(t, 13, cu[t][12])  # last7に値を移動
                            update_user2(t, 12, cu[t][11])  # last6に値を移動
                            update_user2(t, 11, cu[t][10])  # last5に値を移動
                            update_user2(t, 10, cu[t][9])   # last4に値を移動
                            update_user2(t, 9, cu[t][8])    # last3に値を移動
                            update_user2(t, 8, cu[t][7])    # last2に値を移動
                            update_user2(t, 7, now)          # last1を現在時刻に更新
                            break  # ループを終了
                
                if not allow2:  # 子機利用が許可されていない場合
                    try:
                        connected(tag)  # タグが接続された場合に処理を実行
                        return False  # 利用許可がないのでFalseを返す
                    except:  # エラーが発生した場合
                        pass  # エラーを無視
                
                return allow  # 許可状況を返す
            except Exception as e:  # エラーが発生した場合
                print("error: %s" % e)  # エラーメッセージを表示
                print("例外発生。係員にご問い合わせください")  # 注意メッセージを表示
                return None  # Noneを返す
        else:
            print("error: tag isn't Type3Tag")  # タグがType3Tagでない場合のエラーメッセージを表示
            return None  # Noneを返す
    except AttributeError:  # 属性エラーが発生した場合
        print('不正をするなバカもの')  # 警告メッセージを表示
        return None  # Noneを返す

# 子機を実行する関数
def unit_do():
    started = time.time()  # 開始時間を記録
    clf = nfc.ContactlessFrontend('usb')  # NFCリーダーをUSBデバイスとして開く
    result = clf.connect(rdwr={'on-connect': unit})  # 接続処理を開始し、接続時の処理を指定
    clf.close()  # NFCリーダーの接続を閉じる
    return result  # 結果を返す

# NFC IDを取得する関数
def idid(tag):
    newdata = tag.idm.hex()  # タグのIDを16進数形式で取得
    print(newdata)  # IDを表示
    flag = False  # フラグを初期化
    find_list = member_call()  # メンバーリストを取得
    
    for k in range(len(find_list)):  # メンバーリストをループ
        target = ':'  # ターゲット文字列を指定
        idx = find_list[k].find(target)  # ターゲットのインデックスを取得
        key = find_list[k][idx + 1:]  # ID部分を取得
        
        if key == newdata:  # 新しいデータがメンバーのIDと一致する場合
            flag = True  # フラグをTrueに設定
            r = find_list[k][:idx]  # メンバーIDを取得
            update_info(6, r)  # infoテーブルを更新
            return False  # 終了
            
    update_info(6, -2)  # 認識失敗を示す
    return False  # 終了

# タッチ時のハンドラを設定して待機する関数
def id_do():
    started = time.time()  # 開始時間を記録
    clf = nfc.ContactlessFrontend('usb')  # NFCリーダーをUSBデバイスとして開く
    return clf.connect(rdwr={'on-connect': idid}, terminate=partial(afrer, 5, started))  # 接続処理を開始

# メンバーのリストを呼び出す関数
def member_call():
    alignment_user()
    call_list = []  # リストを初期化
    cu = call_user()  # ユーザーデータを取得
    
    for i in range(len(cu)):  # ユーザーデータをループ
        call_list.append(f'{cu[i][0]}:{cu[i][1]}')  # UID:カードIDの形式でリストに追加
    
    return call_list  # 完成したリストを返す

# ユニットのリストを呼び出す関数
def units_call():
    call_list = []  # リストを初期化
    cu = call_units()  # ユニットデータを取得
    
    for i in range(len(cu)):  # ユニットデータをループ
        call_list.append(f'{cu[i][0]}:{cu[i][1]}')  # UID:ユニット名の形式でリストに追加
    
    return call_list  # 完成したリストを返す

# ヒストリーのリストを呼び出す関数
def his_call():
    call_list = []  # リストを初期化
    ch = call_his()  # ヒストリーデータを取得
    
    for i in range(len(ch)):  # ヒストリーデータをループ
        call_list.append(f'{ch[i][0]}:{ch[i][1]}')  # ID:テキストの形式でリストに追加
    
    print(call_list)  # 完成したリストを表示
    return call_list  # 完成したリストを返す

# ユーザー情報を追加的に取得する関数
def add_find():
    ci = call_info()  # infoテーブルからデータを取得
    if ci is None or len(ci) <= 6:
        return "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""  # 空の情報を返す
    
    R = int(ci[6])  # infoテーブルからIDを取得
    if R != -2:  # IDが-2でない場合
        cu = call_user()  # ユーザーデータを取得
        if R < len(cu):  # インデックスが範囲内かチェック
            cu_user = list(cu[R])  # 該当するユーザーの情報を取得
            
            for i in range(len(cu_user)):  # ユーザーの情報をループ
                if cu_user[i] == cu_user[3] and i != 3:  # 特定のユーザーの場合
                    cu_user[i] = "記録なし"  # 情報を置き換え
            
            # ユーザーの情報を取得
            cardid = cu_user[1]
            stock = cu_user[4]
            allow = cu_user[2]
            entry = cu_user[3]
            total = cu_user[6]
            today = cu_user[5]
            last1 = cu_user[7]
            last2 = cu_user[8]
            last3 = cu_user[9]
            last4 = cu_user[10]
            last5 = cu_user[11]
            last6 = cu_user[12]
            last7 = cu_user[13]
            last8 = cu_user[14]
            last9 = cu_user[15]
            last10 = cu_user[16]
            
            return cardid, stock, allow, entry, total, today, last1, last2, last3, last4, last5, last6, last7, last8, last9, last10  # 情報を返す
        else:
            return "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""  # 空の情報を返す
    else:
        return "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""  # 空の情報を返す

# ユニット情報を追加的に取得する関数
def unit_find():
    ci = call_info()  # infoテーブルからデータを取得
    if ci is None or len(ci) <= 6:
        return "", "", "", "", ""  # 空の情報を返す
        
    R = int(ci[6])  # infoテーブルからIDを取得
    if R != -2:  # IDが-2でない場合
        units = call_units()  # ユニットデータを取得
        if R - 1 < len(units):  # インデックスが範囲内かチェック
            cu = list(units[R - 1])  # 指定されたユニットの情報を取得
            unit_name = cu[1]  # ユニット名を取得
            unit_pass = cu[2]  # パスワードを取得
            unit_stock = cu[3]  # 残数を取得
            unit_connect = cu[4]  # 接続数を取得
            unit_available = cu[5]  # 利用可能数を取得
            return unit_name, unit_pass, unit_stock, unit_connect, unit_available  # 取得した情報を返す
        else:
            return "", "", "", "", ""  # 空の情報を返す
    else:
        return "", "", "", "", ""  # 空の情報を返す

# ヒストリーのバックアップを作成する関数
def make_his_backup():
    book = openpyxl.Workbook()  # 新しいExcelワークブックを作成
    sheet = book.active  # アクティブなシートを取得
    ch = call_his()  # ヒストリーデータを取得
    ci = call_info()  # infoデータを取得
    
    # ヒストリーデータをExcelに書き込む
    for i in range(len(ch)):
        sheet[f'A{i + 5}'] = ch[i][0]  # IDを書き込む
        sheet[f'B{i + 5}'] = ch[i][1]  # テキストを書き込む
    
    sheet.protection.password = ci[0]  # パスワードを設定
    sheet.protection.enable()  # シート保護を有効化
    book.save('history.xlsx')  # Excelファイルを保存
    book.close()  # ワークブックを閉じる

# バックアップを作成する関数
def make_backup():
    book = openpyxl.Workbook()  # 新しいExcelワークブックを作成
    sheet = book.active  # アクティブなシートを取得
    cu = call_user()  # ユーザーデータを取得
    ci = call_info()  # infoデータを取得
    
    # ヘッダーを設定
    sheet['A4'] = '各種情報'
    sheet['S4'] = 'id'
    sheet['C4'] = 'カードid'
    sheet['D4'] = '1許可0不可'
    sheet['E4'] = '登録日'
    sheet['F4'] = '残数'
    sheet['G4'] = '今日の利用数'
    sheet['H4'] = '累計利用数'
    sheet['I4'] = '最終利用日'
    sheet['J4'] = '履歴２'
    sheet['K4'] = '履歴３'
    sheet['L4'] = '履歴４'
    sheet['M4'] = '履歴５'
    sheet['N4'] = '履歴６'
    sheet['O4'] = '履歴７'
    sheet['P4'] = '履歴８'
    sheet['Q4'] = '履歴９'
    sheet['R4'] = '履歴１０'
    sheet['B6'] = '日付（日）'
    sheet['B7'] = '増加まで'
    sheet['B8'] = '増加頻度'
    sheet['B9'] = '最大数'
    sheet['B10'] = '←触るな'
    sheet['B11'] = '←触るな'
    
    # ユーザー情報を追加
    for i in range(len(cu)):
        sheet[f'S{i + 5}'] = cu[i][0]  # IDを書き込む
        sheet[f'C{i + 5}'] = cu[i][1]  # カードIDを書き込む
        sheet[f'D{i + 5}'] = cu[i][2]  # 許可状態を書き込む
        sheet[f'E{i + 5}'] = cu[i][3]  # 登録日を書き込む
        sheet[f'F{i + 5}'] = cu[i][4]  # 残数を書き込む
        sheet[f'G{i + 5}'] = cu[i][5]  # 今日の利用数を書き込む
        sheet[f'H{i + 5}'] = cu[i][6]  # 累計利用数を書き込む
        sheet[f'I{i + 5}'] = cu[i][7]  # 最終利用日を書き込む
        sheet[f'J{i + 5}'] = cu[i][8]  # 履歴データを書き込む
        sheet[f'K{i + 5}'] = cu[i][9]  # 履歴データを書き込む
        sheet[f'L{i + 5}'] = cu[i][10]  # 履歴データを書き込む
        sheet[f'M{i + 5}'] = cu[i][11]  # 履歴データを書き込む
        sheet[f'N{i + 5}'] = cu[i][12]  # 履歴データを書き込む
        sheet[f'O{i + 5}'] = cu[i][13]  # 履歴データを書き込む
        sheet[f'P{i + 5}'] = cu[i][14]  # 履歴データを書き込む
        sheet[f'Q{i + 5}'] = cu[i][15]  # 履歴データを書き込む
        sheet[f'R{i + 5}'] = cu[i][16]  # 履歴データを書き込む
    
    # 情報を追加
    for j in range(len(ci)):
        if j != 0:
            sheet[f'A{j + 5}'] = ci[j]  # infoデータを書き込む
    
    sheet.protection.password = ci[0]  # パスワードを設定
    sheet.protection.enable()  # シート保護を有効化
    book.save('データ復元時はファイル名を「backup」にしてください.xlsx')  # Excelファイルを保存
    book.close()  # ワークブックを閉じる

# Excelからデータを復元する関数
def copy_from_excel(txt):
    book = openpyxl.load_workbook('backup.xlsx')  # バックアップファイルをロード
    sheet = book.active  # アクティブなシートを取得
    conn = create_server_connection("localhost", "root", "Hiramekigo@1", "userdb")  # データベースに接続
    cur = conn.cursor(buffered=True)  # カーソルを作成
    cur.execute('TRUNCATE user')  # ユーザーテーブルを空にする
    cur.execute('TRUNCATE info')  # インフォテーブルを空にする
    user_info = None  # ユーザー情報を初期化
    num = 5  # データ開始行番号を指定
    
    while True:  # データ読み込みループ
        if sheet[f'E{num}'].value is None:  # 登録日がNoneの場合
            break  # ループを終了
        idnum = sheet[f'S{num}'].value  # IDを取得
        cardid = sheet[f'C{num}'].value  # カードIDを取得
        allow = sheet[f'D{num}'].value  # 許可状態を取得
        entry = sheet[f'E{num}'].value  # 登録日を取得
        stock = sheet[f'F{num}'].value  # 残数を取得
        today = sheet[f'G{num}'].value  # 今日の利用数を取得
        total = sheet[f'H{num}'].value  # 累計利用数を取得
        last1 = sheet[f'I{num}'].value  # 履歴データを取得
        last2 = sheet[f'J{num}'].value  # 履歴データを取得
        last3 = sheet[f'K{num}'].value  # 履歴データを取得
        last4 = sheet[f'L{num}'].value  # 履歴データを取得
        last5 = sheet[f'M{num}'].value  # 履歴データを取得
        last6 = sheet[f'N{num}'].value  # 履歴データを取得
        last7 = sheet[f'O{num}'].value  # 履歴データを取得
        last8 = sheet[f'P{num}'].value  # 履歴データを取得
        last9 = sheet[f'Q{num}'].value  # 履歴データを取得
        last10 = sheet[f'R{num}'].value  # 履歴データを取得
        
        user_info = idnum, cardid, allow, entry, stock, today, total, last1, last2, last3, last4, last5, last6, last7, last8, last9, last10  # ユーザー情報をタプルにまとめ
        cur.execute(f'INSERT INTO user(id, cardid, allow, entry, stock, today, total, last1, last2, last3, last4, last5, last6, last7, last8, last9, last10) VALUES({idnum}, "{cardid}", {allow}, "{entry}", {stock}, {today}, {total}, "{last1}", "{last2}", "{last3}", "{last4}", "{last5}", "{last6}", "{last7}", "{last8}", "{last9}", "{last10}")')  # ユーザー情報をデータベースに挿入
        num += 1  # 行番号をインクリメント
    
    # データリストを作成
    data_list = []
    for num in range(5, 12, 1):  # データを5行目から11行目まで取得
        mydata = sheet[f'A{num}'].value  # データを取得
        if num == 5:  # 5行目の場合
            mydata = txt  # テキストを設定
        data_list.append(mydata)  # データリストに追加
    
    # infoテーブルにデータを挿入
    cur.execute(f'INSERT INTO info (pass, daycount, updatecount, freq, maximum, done, list) VALUES("{data_list[0]}", {data_list[1]}, {data_list[2]}, {data_list[3]}, {data_list[4]}, {data_list[5]}, "{data_list[6]}")')
    conn.commit()  # 変更をコミット    
    cur.close()  # カーソルを閉じる
    conn.close()  # 接続を閉じる

# 管理者パスワードを更新する関数
def update_admin_password(new_password):
    """管理者パスワードを更新する関数"""
    conn = create_server_connection("localhost", "root", "Hiramekigo@1", "userdb")
    if conn is not None:
        cur = conn.cursor()
        cur.execute("UPDATE info SET pass = %s", (new_password,))
        conn.commit()
        cur.close()
        conn.close()
        print(f"管理者パスワードを '{new_password}' に更新しました。")
    else:
        print("データベースに接続できませんでした。")

def check_reader_status():
    """True=接続済み / False=未接続"""
    try:
        import nfc
        clf = nfc.ContactlessFrontend("usb")
        if clf is None:
            return False
        clf.close()
        return True
    except Exception:
        return False

def get_db_connection():
    """データベース接続を取得する関数"""
    conn = mysql.connector.connect(
        host="localhost",
        user="root",               # 必要に応じて変更
        password="Hiramekigo@1",   # MySQL のパスワード
        database="userdb"
    )
    return conn
